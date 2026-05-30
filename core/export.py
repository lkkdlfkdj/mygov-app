"""
导出工具模块 ExportUtils
- Excel（带样式） — 使用 pandas + openpyxl
- PDF（中文不乱码） — 使用 FPDF2 + 内置中文字体
- CSV
- JSON
- TXT
- 文件保存在 data/exports/ 目录
"""

import os
import json
import csv
from datetime import datetime

# 导出的基础目录
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXPORT_DIR = os.path.join(BASE_DIR, 'data', 'exports')
FONT_DIR = os.path.join(BASE_DIR, 'assets', 'fonts')

# 中文字体路径（用于PDF导出）
FONT_PATH = os.path.join(FONT_DIR, 'simhei.ttf')
FONT_FALLBACK = os.path.join(FONT_DIR, 'msyh.ttc')


# ==================== 列定义 ====================
# 每种表的导出列（显示名称 -> 字段名）

COMPLAINT_COLUMNS = [
    ('编号', 'id'),
    ('标题', 'title'),
    ('内容', 'content'),
    ('投诉人', 'complainant'),
    ('电话', 'phone'),
    ('地址', 'address'),
    ('状态', 'status'),
    ('紧急程度', 'urgency'),
    ('办理期限', 'deadline'),
    ('回复内容', 'reply'),
    ('提交时间', 'created_at'),
]

HAZARD_COLUMNS = [
    ('编号', 'id'),
    ('位置', 'location'),
    ('纬度', 'latitude'),
    ('经度', 'longitude'),
    ('权属类别', 'ownership_type'),
    ('隐患类型', 'hazard_type'),
    ('备注', 'remark'),
    ('状态', 'status'),
    ('上报时间', 'created_at'),
]

CASE_COLUMNS = [
    ('编号', 'id'),
    ('案件类型', 'case_type'),
    ('案件编号', 'case_number'),
    ('编号起始', 'number_range_start'),
    ('编号结束', 'number_range_end'),
    ('当事人', 'party_name'),
    ('违法事实', 'violation_fact'),
    ('状态', 'status'),
    ('创建时间', 'created_at'),
]

LAW_COLUMNS = [
    ('编号', 'id'),
    ('类别', 'category'),
    ('标题', 'title'),
    ('违法行为', 'violation'),
    ('禁止规定', 'prohibition'),
    ('处罚依据', 'penalty_law'),
    ('处罚标准', 'penalty_standard'),
]

AD_COLUMNS = [
    ('编号', 'id'),
    ('申请人', 'applicant'),
    ('店铺名称', 'shop_name'),
    ('地址', 'address'),
    ('查勘状态', 'survey_status'),
    ('审批状态', 'approval_status'),
    ('申请时间', 'created_at'),
]

# 表名 -> (列定义, 数据获取方法名)
TABLE_CONFIG = {
    'complaints': (COMPLAINT_COLUMNS, 'get_all_complaints'),
    'hazards': (HAZARD_COLUMNS, 'get_all_hazards'),
    'cases': (CASE_COLUMNS, 'get_all_cases'),
    'laws': (LAW_COLUMNS, 'get_all_laws'),
    'ads': (AD_COLUMNS, 'get_all_ads'),
}


def _ensure_export_dir():
    """确保导出目录存在"""
    os.makedirs(EXPORT_DIR, exist_ok=True)
    return EXPORT_DIR


def _get_table_data(table_name):
    """从Storage获取指定表的数据和列定义"""
    from core.storage import Storage
    storage = Storage()

    if table_name not in TABLE_CONFIG:
        return None, None, f'未知表名: {table_name}'

    columns, method_name = TABLE_CONFIG[table_name]

    # 调用对应的Storage方法获取数据
    method = getattr(storage, method_name)
    rows = method()

    return columns, rows, None


def _format_value(value, max_len=None):
    """格式化单元格值"""
    if value is None:
        return ''
    s = str(value)
    if max_len and len(s) > max_len:
        s = s[:max_len - 3] + '...'
    return s


# ==================== Excel 导出 ====================

def export_excel(table_name, filename=None):
    """
    导出为Excel文件（带样式）
    返回: (success, filepath_or_error)
    """
    columns, rows, error = _get_table_data(table_name)
    if error:
        return False, error

    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        _ensure_export_dir()

        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'{table_name}_{timestamp}.xlsx'

        filepath = os.path.join(EXPORT_DIR, filename)

        # 准备数据
        header_labels = [c[0] for c in columns]
        field_names = [c[1] for c in columns]

        data_rows = []
        for row in rows:
            data_row = []
            for field in field_names:
                val = row.get(field, '')
                if isinstance(val, float):
                    val = round(val, 6)
                data_row.append('' if val is None else val)
            data_rows.append(data_row)

        # 用 pandas 创建 DataFrame
        df = pd.DataFrame(data_rows, columns=header_labels)

        # 写入 Excel
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=table_name)
            workbook = writer.book
            worksheet = writer.sheets[table_name]

            # ---- 样式定义 ----
            header_font = Font(
                name='微软雅黑',
                bold=True,
                color='FFFFFF',
                size=11,
            )
            header_fill = PatternFill(
                start_color='2E7D32',
                end_color='2E7D32',
                fill_type='solid'
            )
            header_alignment = Alignment(
                horizontal='center',
                vertical='center',
            )
            cell_font = Font(name='微软雅黑', size=10)
            cell_alignment = Alignment(
                horizontal='left',
                vertical='center',
                wrap_text=True,
            )
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC'),
            )

            # ---- 设置表头样式 ----
            for col_idx, header in enumerate(header_labels, start=1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # ---- 设置数据单元格样式 ----
            for row_idx in range(2, len(data_rows) + 2):
                for col_idx in range(1, len(header_labels) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = cell_font
                    cell.alignment = cell_alignment
                    cell.border = thin_border

            # ---- 自动调整列宽 ----
            for col_idx, header in enumerate(header_labels, start=1):
                # 计算该列最大字符宽度（中文字符按2算）
                max_width = len(header) * 2  # 表头宽度
                for row_idx in range(2, min(len(data_rows) + 2, 50)):  # 只采样前50行
                    cell_value = str(worksheet.cell(row=row_idx, column=col_idx).value or '')
                    # 估算宽度：中文字符2，英文1
                    char_width = sum(2 if ord(c) > 127 else 1 for c in cell_value)
                    max_width = max(max_width, min(char_width, 60))

                # 限制列宽在8~40之间
                col_width = max(8, min(max_width + 2, 40))
                worksheet.column_dimensions[get_column_letter(col_idx)].width = col_width

            # ---- 冻结首行 ----
            worksheet.freeze_panes = 'A2'

        return True, filepath

    except Exception as e:
        return False, f'Excel导出失败: {str(e)}'


# ==================== PDF 导出 ====================

def export_pdf(table_name, filename=None):
    """
    导出为PDF文件（中文不乱码）
    使用 FPDF2 + 内置 simhei.ttf 中文字体
    返回: (success, filepath_or_error)
    """
    columns, rows, error = _get_table_data(table_name)
    if error:
        return False, error

    try:
        from fpdf import FPDF

        _ensure_export_dir()

        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'{table_name}_{timestamp}.pdf'

        filepath = os.path.join(EXPORT_DIR, filename)

        header_labels = [c[0] for c in columns]
        field_names = [c[1] for c in columns]

        # ---- 创建PDF ----
        pdf = FPDF(orientation='L', unit='mm', format='A4')
        pdf.set_auto_page_break(auto=True, margin=15)

        # 注册中文字体
        if os.path.exists(FONT_PATH):
            pdf.add_font('Chinese', '', FONT_PATH, uni=True)
            font_name = 'Chinese'
        elif os.path.exists(FONT_FALLBACK):
            pdf.add_font('Chinese', '', FONT_FALLBACK, uni=True)
            font_name = 'Chinese'
        else:
            # 无中文字体时使用默认字体（中文会显示为方框）
            font_name = 'Helvetica'

        # ---- 计算列宽 ----
        page_width = pdf.w - 20  # 左右各10mm边距
        col_count = len(header_labels)
        # 根据列数动态分配列宽
        if col_count <= 5:
            col_widths = [page_width / col_count] * col_count
        else:
            # 固定最小宽度，剩余均分
            min_width = 20
            remaining = page_width - min_width * col_count
            if remaining < 0:
                col_widths = [page_width / col_count] * col_count
            else:
                col_widths = [min_width + remaining / col_count] * col_count

        # ---- 表头 ----
        pdf.add_page()
        pdf.set_font(font_name, '', 14)
        title_text = {
            'complaints': '投诉管理数据导出',
            'hazards': '隐患上报数据导出',
            'cases': '案件采集数据导出',
            'laws': '法条库数据导出',
            'ads': '店招申请数据导出',
        }.get(table_name, f'{table_name}数据导出')
        pdf.cell(0, 10, title_text, new_x='LMARGIN', new_y='NEXT', align='C')
        pdf.ln(5)

        # 导出时间
        pdf.set_font(font_name, '', 8)
        pdf.cell(0, 5, f'导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}',
                 new_x='LMARGIN', new_y='NEXT', align='R')
        pdf.ln(3)

        # ---- 绘制表格 ----
        # 表头背景色
        pdf.set_fill_color(46, 125, 50)  # 深绿
        pdf.set_text_color(255, 255, 255)
        pdf.set_font(font_name, '', 9)

        for i, header in enumerate(header_labels):
            pdf.cell(col_widths[i], 8, header, border=1, fill=True, align='C')
        pdf.ln()

        # 数据行
        pdf.set_text_color(33, 33, 33)
        pdf.set_font(font_name, '', 8)

        # 交替行颜色
        light_gray = (245, 245, 245)
        white = (255, 255, 255)

        for row_idx, row in enumerate(rows):
            # 检查是否需要翻页
            if pdf.y > 260:
                pdf.add_page()
                # 重新打印表头
                pdf.set_fill_color(46, 125, 50)
                pdf.set_text_color(255, 255, 255)
                pdf.set_font(font_name, '', 9)
                for i, header in enumerate(header_labels):
                    pdf.cell(col_widths[i], 8, header, border=1, fill=True, align='C')
                pdf.ln()
                pdf.set_text_color(33, 33, 33)
                pdf.set_font(font_name, '', 8)

            # 行背景色
            bg = light_gray if row_idx % 2 == 0 else white
            pdf.set_fill_color(*bg)

            # 限制行高
            row_height = 7

            for col_idx, field in enumerate(field_names):
                val = _format_value(row.get(field, ''), max_len=80)
                pdf.cell(col_widths[col_idx], row_height, val,
                         border=1, fill=True, align='L')

            pdf.ln()

        # ---- 页脚总结 ----
        pdf.ln(5)
        pdf.set_font(font_name, '', 9)
        pdf.cell(0, 5, f'共 {len(rows)} 条记录',
                 new_x='LMARGIN', new_y='NEXT', align='C')

        pdf.output(filepath)
        return True, filepath

    except Exception as e:
        return False, f'PDF导出失败: {str(e)}'


# ==================== CSV 导出 ====================

def export_csv(table_name, filename=None):
    """
    导出为CSV文件（标准CSV格式，UTF-8 BOM）
    返回: (success, filepath_or_error)
    """
    columns, rows, error = _get_table_data(table_name)
    if error:
        return False, error

    try:
        _ensure_export_dir()

        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'{table_name}_{timestamp}.csv'

        filepath = os.path.join(EXPORT_DIR, filename)

        header_labels = [c[0] for c in columns]
        field_names = [c[1] for c in columns]

        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            # 写表头
            writer.writerow(header_labels)
            # 写数据
            for row in rows:
                values = []
                for field in field_names:
                    val = row.get(field, '')
                    if val is None:
                        val = ''
                    # 确保数字不会丢失精度
                    if isinstance(val, float):
                        val = round(val, 6)
                    values.append(str(val))
                writer.writerow(values)

        return True, filepath

    except Exception as e:
        return False, f'CSV导出失败: {str(e)}'


# ==================== JSON 导出 ====================

def export_json(table_name, filename=None):
    """
    导出为JSON文件（格式化，含导出时间）
    返回: (success, filepath_or_error)
    """
    rows, error = _get_data(table_name)
    if error:
        return False, error

    try:
        _ensure_export_dir()

        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'{table_name}_{timestamp}.json'

        filepath = os.path.join(EXPORT_DIR, filename)

        # 对 rouned 浮点数进行格式化
        clean_rows = []
        for row in rows:
            clean_row = {}
            for k, v in row.items():
                if isinstance(v, float):
                    clean_row[k] = round(v, 6)
                elif isinstance(v, (list, dict)):
                    clean_row[k] = str(v)
                else:
                    clean_row[k] = v
            clean_rows.append(clean_row)

        output = {
            'table': table_name,
            'export_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'total_count': len(clean_rows),
            'data': clean_rows,
        }

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(output, f, ensure_ascii=False, indent=2)

        return True, filepath

    except Exception as e:
        return False, f'JSON导出失败: {str(e)}'


def _get_data(table_name):
    """与 _get_table_data 类似但不返回列定义"""
    from core.storage import Storage
    storage = Storage()

    if table_name not in TABLE_CONFIG:
        return None, f'未知表名: {table_name}'

    _, method_name = TABLE_CONFIG[table_name]
    method = getattr(storage, method_name)
    rows = method()

    return rows, None


# ==================== TXT 导出 ====================

def export_txt(table_name, filename=None):
    """
    导出为TXT文件（纯文本报告格式）
    返回: (success, filepath_or_error)
    """
    columns, rows, error = _get_table_data(table_name)
    if error:
        return False, error

    try:
        _ensure_export_dir()

        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'{table_name}_{timestamp}.txt'

        filepath = os.path.join(EXPORT_DIR, filename)

        header_labels = [c[0] for c in columns]
        field_names = [c[1] for c in columns]

        title_text = {
            'complaints': '投诉管理数据报告',
            'hazards': '隐患上报数据报告',
            'cases': '案件采集数据报告',
            'laws': '法条库数据报告',
            'ads': '店招申请数据报告',
        }.get(table_name, f'{table_name}数据报告')

        lines = []
        # 分隔线
        sep = '=' * 60
        line_sep = '-' * 60

        lines.append(sep)
        lines.append(f'  {title_text}')
        lines.append(sep)
        lines.append(f'  导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
        lines.append(f'  记录总数: {len(rows)}')
        lines.append(sep)
        lines.append('')

        for idx, row in enumerate(rows, start=1):
            lines.append(f'  【记录 {idx}】')
            lines.append(line_sep)
            for col_idx, field in enumerate(field_names):
                val = _format_value(row.get(field, ''))
                label = header_labels[col_idx]
                lines.append(f'    {label}: {val}')
            lines.append(line_sep)
            lines.append('')

        lines.append(sep)
        lines.append(f'  共 {len(rows)} 条记录')
        lines.append(f'  导出完成: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
        lines.append(sep)

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))

        return True, filepath

    except Exception as e:
        return False, f'TXT导出失败: {str(e)}'


# ==================== 统一导出接口 ====================

EXPORT_FORMATS = {
    'excel': {
        'label': 'Excel (.xlsx)',
        'extension': '.xlsx',
        'func': export_excel,
    },
    'pdf': {
        'label': 'PDF (.pdf)',
        'extension': '.pdf',
        'func': export_pdf,
    },
    'csv': {
        'label': 'CSV (.csv)',
        'extension': '.csv',
        'func': export_csv,
    },
    'json': {
        'label': 'JSON (.json)',
        'extension': '.json',
        'func': export_json,
    },
    'txt': {
        'label': 'TXT (.txt)',
        'extension': '.txt',
        'func': export_txt,
    },
}


def export_data(table_name, export_format='excel', filename=None):
    """
    统一导出接口

    参数:
        table_name: 'complaints' | 'hazards' | 'cases' | 'laws' | 'ads'
        export_format: 'excel' | 'pdf' | 'csv' | 'json' | 'txt'
        filename: 自定义文件名（可选）

    返回:
        (success: bool, message_or_filepath: str)

    示例:
        success, result = export_data('complaints', 'excel')
        if success:
            print(f'文件已保存: {result}')
        else:
            print(f'导出失败: {result}')
    """
    if export_format not in EXPORT_FORMATS:
        return False, f'不支持的导出格式: {export_format}，可选: {", ".join(EXPORT_FORMATS.keys())}'

    if table_name not in TABLE_CONFIG:
        return False, f'不支持的导出表名: {table_name}，可选: {", ".join(TABLE_CONFIG.keys())}'

    format_info = EXPORT_FORMATS[export_format]
    func = format_info['func']

    return func(table_name, filename)


def get_export_dir():
    """获取导出目录路径"""
    _ensure_export_dir()
    return EXPORT_DIR


def clean_exports(days=30):
    """
    清理超过指定天数的导出文件
    返回: (deleted_count, error)
    """
    _ensure_export_dir()
    now = datetime.now()
    deleted = 0

    for fname in os.listdir(EXPORT_DIR):
        fpath = os.path.join(EXPORT_DIR, fname)
        if os.path.isfile(fpath):
            mtime = datetime.fromtimestamp(os.path.getmtime(fpath))
            if (now - mtime).days > days:
                try:
                    os.remove(fpath)
                    deleted += 1
                except Exception:
                    pass

    return deleted, None


# ==================== 全局暴露 ====================

__all__ = [
    'export_data',
    'export_excel',
    'export_pdf',
    'export_csv',
    'export_json',
    'export_txt',
    'get_export_dir',
    'clean_exports',
    'EXPORT_FORMATS',
    'TABLE_CONFIG',
    'EXPORT_DIR',
]
